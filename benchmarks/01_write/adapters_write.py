import openpyxl
import xlsxwriter
import pyexcelerate
import fastxlsx
import pylightxl
import numpy as np
from typing import Any
from base_write import BaseWorkbook
from fastxlsx import RangeInfo, DShape, DType
from typing import List, Tuple


class FastXLSXWorkbook(BaseWorkbook):
    def __init__(self):
        self.wb = fastxlsx.WriteOnlyWorkbook()

    def create_sheet(self, title: str) -> Any:
        return self.wb.create_sheet(title)

    def write_func(self, ws: Any, row: int, col: int, value: Any, dtype: DType) -> None:
        pass

    def write_to_sheet(
        self,
        ws: fastxlsx.WriteOnlyWorksheet,
        data_to_write: List[Tuple[RangeInfo, Any]],
    ) -> None:
        def to_array(value, dtype: DType):
            if dtype in [DType.Bool, DType.Int, DType.Float]:
                return np.asarray(value)
            else:
                return value

        for range_info, value in data_to_write:
            dtype = range_info.dtype
            if isinstance(range_info.data_shape, DShape.Scalar):
                ws.write_cell(range_info.pos, value, dtype=dtype)
            elif isinstance(range_info.data_shape, DShape.Row):
                ws.write_row(range_info.pos, to_array(value, dtype), dtype=dtype)
            elif isinstance(range_info.data_shape, DShape.Column):
                ws.write_column(range_info.pos, to_array(value, dtype), dtype=dtype)
            elif isinstance(range_info.data_shape, DShape.Matrix):
                ws.write_matrix(range_info.pos, to_array(value, dtype), dtype=dtype)

    def save(self, filename: str) -> None:
        self.wb.save(filename)


class FastXLSXCellwiseWorkbook(BaseWorkbook):
    def __init__(self):
        self.wb = fastxlsx.WriteOnlyWorkbook()

    def create_sheet(self, title: str) -> Any:
        return self.wb.create_sheet(title)

    def write_func(self, ws: Any, row: int, col: int, value: Any, dtype: DType) -> None:
        ws.write_cell((row, col), value, dtype=dtype)

    def save(self, filename: str) -> None:
        self.wb.save(filename)


class OpenPyXLWorkbook(BaseWorkbook):
    def __init__(self):
        self.wb = openpyxl.Workbook(write_only=False)
        self.wb.remove(self.wb.active)

    def create_sheet(self, title: str) -> Any:
        return self.wb.create_sheet(title)

    def write_func(self, ws: Any, row: int, col: int, value: Any, dtype: DType) -> None:
        ws.cell(row + 1, col + 1, value)

    def save(self, filename: str) -> None:
        self.wb.save(filename)


class OpenPyXLWriteonlyWorkbook(BaseWorkbook):
    def __init__(self):
        self.wb = openpyxl.Workbook(write_only=True)

    def create_sheet(self, title: str) -> Any:
        return self.wb.create_sheet(title)

    def write_func(
        self, data: List[List[Any]], row: int, col: int, value: Any, dtype: DType
    ) -> None:
        data[row][col] = value

    def write_to_sheet(
        self, ws: Any, data_to_write: List[Tuple[RangeInfo, Any]]
    ) -> None:
        end_list = [range_info.end for range_info, _ in data_to_write]
        n_rows, n_cols = tuple(map(lambda x: max(x) + 1, zip(*end_list)))
        data = [[None for _ in range(n_cols)] for _ in range(n_rows)]
        super().write_to_sheet(data, data_to_write)
        for row in data:
            ws.append(row)

    def save(self, filename: str) -> None:
        self.wb.save(filename)


class XlsxWriterWorkbook(BaseWorkbook):
    def __init__(self):
        self.wb = xlsxwriter.Workbook()

    def create_sheet(self, title: str) -> Any:
        return self.wb.add_worksheet(title)

    def write_func(self, ws: Any, row: int, col: int, value: Any, dtype: DType) -> None:
        ws.write(row, col, value)

    def save(self, filename: str) -> None:
        self.wb.filename = filename
        self.wb.close()


class PyExcelerateWorkbook(BaseWorkbook):
    def __init__(self):
        self.wb = pyexcelerate.Workbook()

    def create_sheet(self, title: str) -> Any:
        return self.wb.new_sheet(title)

    def write_func(self, ws: Any, row: int, col: int, value: Any, dtype: DType) -> None:
        ws.set_cell_value(row, col, value)

    def save(self, filename: str) -> None:
        self.wb.save(filename)


class PyLightXLWorkbook(BaseWorkbook):
    def __init__(self):
        self.wb = pylightxl.Database()

    def create_sheet(self, title: str) -> Any:
        self.wb.add_ws(ws=title)
        return self.wb.ws(ws=title)

    def write_func(self, ws: Any, row: int, col: int, value: Any, dtype: DType) -> None:
        ws.update_index(row=row + 1, col=col + 1, val=value)

    def save(self, filename: str) -> None:
        pylightxl.writexl(self.wb, filename)
