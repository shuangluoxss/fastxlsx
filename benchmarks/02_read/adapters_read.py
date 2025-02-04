import openpyxl
import pylightxl
import fastxlsx
import python_calamine
import numpy as np
from typing import Any
from base_read import BaseWorkbook
from fastxlsx import RangeInfo, DShape, DType, idx_to_addr
from typing import List, Tuple, Union, Dict


class FastXLSXWorkbook(BaseWorkbook):
    def __init__(self, filename):
        self.wb = fastxlsx.ReadOnlyWorkbook(filename)

    def get_sheet(self, idx_or_name: Union[int, str]) -> Any:
        return self.wb.get(idx_or_name)

    def read_func(self, ws: Any, row: int, col: int, dtype: DType) -> Any:
        pass

    def read_from_sheet(
        self,
        ws: fastxlsx.ReadOnlyWorksheet,
        range_to_read: Dict[str, RangeInfo],
    ) -> None:
        return ws.read_values(range_to_read)


class FastXLSXCellwiseWorkbook(BaseWorkbook):
    def __init__(self, filename):
        self.wb = fastxlsx.ReadOnlyWorkbook(filename)

    def get_sheet(self, idx_or_name: Union[int, str]) -> Any:
        return self.wb.get(idx_or_name)

    def read_func(self, ws: Any, row: int, col: int, dtype: DType) -> Any:
        return ws.cell_value((row, col), dtype=dtype)


class OpenPyXLWorkbook(BaseWorkbook):
    def __init__(self, filename):
        self.wb = openpyxl.load_workbook(filename)

    def get_sheet(self, idx_or_name: Union[int, str]) -> Any:
        if isinstance(idx_or_name, int):
            return self.wb.worksheets[idx_or_name]
        else:
            return self.wb[idx_or_name]

    def read_func(self, ws: Any, row: int, col: int, dtype: DType) -> Any:
        return ws.cell(row + 1, col + 1).value


class PyLightXLWorkbook(BaseWorkbook):
    def __init__(self, filename):
        self.wb = pylightxl.readxl(filename)

    def get_sheet(self, idx_or_name: Union[int, str]) -> Any:
        if isinstance(idx_or_name, int):
            return self.wb.ws(self.wb.ws_names[idx_or_name])
        else:
            return self.wb.ws(idx_or_name)

    def read_func(self, ws: Any, row: int, col: int, dtype: DType) -> Any:
        return ws.index(row + 1, col + 1)


class PyCalamineWorkbook(BaseWorkbook):
    def __init__(self, filename):
        self.wb = python_calamine.load_workbook(filename)

    def get_sheet(self, idx_or_name: Union[int, str]) -> Any:
        if isinstance(idx_or_name, int):
            return self.wb.get_sheet_by_index(idx_or_name)
        else:
            return self.wb.get_sheet_by_name(idx_or_name)

    def read_func(self, data: Any, row: int, col: int, dtype: DType) -> Any:
        return data[row][col]

    def read_from_sheet(
        self,
        ws: python_calamine.CalamineSheet,
        range_to_read: Dict[str, RangeInfo],
    ) -> None:
        data = ws.to_python()
        return super().read_from_sheet(data, range_to_read)
